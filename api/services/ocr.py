"""
ocr.py — Text extraction service for all supported file types.

Extraction strategy per format:
  .pdf   → PyMuPDF (fast text-layer extraction), fallback to Tesseract OCR
  .png / .jpg / .jpeg → Tesseract OCR via pytesseract
  .docx  → python-docx paragraph extraction
  .txt / .md / .csv → plain UTF-8 read
  .xlsx  → openpyxl cell-value extraction
"""

from pathlib import Path
from typing import Optional
import traceback

from rich.console import Console

console = Console(stderr=True)

# Maximum characters of extracted text we store per file
MAX_TEXT_LENGTH = 10_000


def extract_text(file_path: Path) -> str:
    """
    Extract readable text from a file using the appropriate method for its extension.

    Args:
        file_path: Absolute path to the file.

    Returns:
        Extracted text string (may be empty if extraction fails or file has no text).
    """
    ext = file_path.suffix.lower()

    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".txt": _extract_plain,
        ".md": _extract_plain,
        ".csv": _extract_plain,
        ".xlsx": _extract_xlsx,
        ".png": _extract_image_ocr,
        ".jpg": _extract_image_ocr,
        ".jpeg": _extract_image_ocr,
    }

    extractor = extractors.get(ext)
    if extractor is None:
        console.print(f"[yellow]Warning:[/yellow] No extractor for extension '{ext}', skipping text extraction.")
        return ""

    try:
        text = extractor(file_path)
        # Truncate to avoid storing huge blobs in the DB
        return text[:MAX_TEXT_LENGTH].strip()
    except Exception as e:
        console.print(f"[yellow]Warning:[/yellow] Text extraction failed for {file_path.name}: {e}")
        return ""


def _extract_pdf(file_path: Path) -> str:
    """
    Extract text from a PDF using PyMuPDF.

    If the PDF has no text layer (e.g., a scanned document), falls back to
    rendering each page as an image and running Tesseract OCR on it.

    Args:
        file_path: Path to the PDF file.

    Returns:
        Extracted text string.
    """
    import fitz  # PyMuPDF

    text_parts: list[str] = []

    with fitz.open(str(file_path)) as doc:
        for page in doc:
            page_text = page.get_text("text").strip()
            if page_text:
                text_parts.append(page_text)

    if text_parts:
        return "\n".join(text_parts)

    # No text layer found — fall back to OCR
    console.print(f"[dim]  PDF has no text layer, using OCR for {file_path.name}[/dim]")
    return _pdf_ocr_fallback(file_path)


def _pdf_ocr_fallback(file_path: Path) -> str:
    """
    Render each PDF page to an image and run Tesseract OCR on it.

    Args:
        file_path: Path to the PDF file.

    Returns:
        OCR-extracted text string.
    """
    import fitz
    import pytesseract
    from PIL import Image
    import io

    text_parts: list[str] = []

    with fitz.open(str(file_path)) as doc:
        for page in doc:
            # Render at 2x scale for better OCR accuracy
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat)
            img_bytes = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_bytes))
            page_text = pytesseract.image_to_string(img)
            if page_text.strip():
                text_parts.append(page_text.strip())

    return "\n".join(text_parts)


def _extract_image_ocr(file_path: Path) -> str:
    """
    Run Tesseract OCR on an image file (PNG, JPG, JPEG).

    Args:
        file_path: Path to the image file.

    Returns:
        OCR-extracted text string.
    """
    import pytesseract
    from PIL import Image

    img = Image.open(str(file_path))
    return pytesseract.image_to_string(img)


def _extract_docx(file_path: Path) -> str:
    """
    Extract text from a .docx file using python-docx.

    Concatenates all paragraph texts, preserving line breaks between them.

    Args:
        file_path: Path to the .docx file.

    Returns:
        Extracted text string.
    """
    from docx import Document

    doc = Document(str(file_path))
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n".join(paragraphs)


def _extract_plain(file_path: Path) -> str:
    """
    Read a plain-text file (.txt, .md, .csv) as UTF-8.

    Falls back to latin-1 encoding if UTF-8 decoding fails.

    Args:
        file_path: Path to the text file.

    Returns:
        File contents as a string.
    """
    try:
        return file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return file_path.read_text(encoding="latin-1")


def _extract_xlsx(file_path: Path) -> str:
    """
    Extract cell values from an Excel workbook (.xlsx) using openpyxl.

    Iterates over all sheets and all cells, joining non-empty values with spaces.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Space-separated string of all cell values across all sheets.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    parts.append(str(cell_val))

    wb.close()
    return " ".join(parts)
